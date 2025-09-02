import os.path
import numpy as np
import pandas as pd

from openpyxl.styles import Font
from openpyxl.styles import Alignment

from entrada import              Entrada
from procesadorJuiciosHelper    import getCompetenciasNoTecnicas, numeroDeOrden, color_rows, ajustarFormatoCeldas, getInstructorEnReporte
from config                     import ancho_columnas, COLUMNAS_DATOS, COLUMNAS_HOJA

class ProcesadorJuicios:
    def __init__(self, folder: str, archivo: str, novedades: pd.DataFrame = None, activos: pd.DataFrame = None, instructores: pd.DataFrame = None):
        self.folder         = folder
        self.archivo        = archivo
        self.df_novedades   = novedades
        self.df_activos     = activos
        self.df_instructores= instructores

        self.ficha              : str = None
        self.codigo_programa    : str = None
        self.version_programa   : str = None
        self.fecha_inicio       : pd.Timestamp = None
        self.fecha_fin          : pd.Timestamp = None

        self.df_hoja                : pd.DataFrame = None
        self.df_datos               : pd.DataFrame = None
        self.df_novedades_ficha     : pd.DataFrame = None
        self.df_activos_ficha       : pd.DataFrame = None
        self.df_instructores_ficha  : pd.DataFrame = None

    def escribirArchivoSalida(self):
        nuevo_nombre = f"{self.ficha}.xlsx"
        try:
            with pd.ExcelWriter(os.path.join(self.folder, nuevo_nombre), engine="openpyxl") as writer:
                self.df_datos.sort_values(by="orden").style.apply(color_rows, axis=1).to_excel(writer, sheet_name="Datos", index=False)  
                workbook = writer.book
                worksheet = workbook['Datos']
                worksheet.delete_cols(11)
                ajustarFormatoCeldas(worksheet, 24, ancho_columnas['datos'])
                # ajustar la fila de nombres de los instructores
                row_dimension = worksheet.row_dimensions[2]
                row_dimension.height = 60
                for cell in worksheet[2]:
                    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')                   

                if not self.df_novedades_ficha is None:
                    self.df_novedades_ficha.to_excel( writer, sheet_name="Novedades", index=False)
                    ajustarFormatoCeldas(workbook['Novedades'], 3, ancho_columnas['novedades'])

                if not self.df_activos_ficha is None:
                    self.df_activos_ficha.to_excel(writer, sheet_name="Activos", index=False)
                    ajustarFormatoCeldas(workbook['Activos'], 3, ancho_columnas['novedades'])

                self.df_hoja.to_excel( writer, sheet_name="Hoja", index=False)
                ajustarFormatoCeldas(workbook['Hoja'], 11, ancho_columnas['hoja'])
                worksheet = workbook['Hoja']
                for fila in range(2,13):
                    worksheet.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
                    worksheet.cell(row=fila, column=1).alignment
                for fila in range(2,13):
                    worksheet.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
                    worksheet.cell(row=fila, column=1).alignment = Alignment(horizontal='left', vertical='center')
                    worksheet.cell(row=fila, column=3).alignment = Alignment(horizontal='left', vertical='center')
                for row in worksheet.iter_rows():
                    row[5].alignment = Alignment(horizontal='left')  
                    row[6].alignment = Alignment(horizontal='left')
                for col in range(1,12):
                    worksheet.cell(1, col).value = None
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
                worksheet.cell(1,1).value = "Reporte de Juicios"
                worksheet.cell(1,1).font = Font(name='Arial', size=20)   

        except Exception as e:
            raise Exception(f"Error: no es posible guardar el archivo: {e}")

    def armar_df_datos(self):
        competencias_no_tecnicas = getCompetenciasNoTecnicas('631101')

        # 1. creamos los dataframes novedades, acitvos, instrucrtroes de la ficha
        if not self.df_novedades is None:
            self.df_novedades_ficha     = self.df_novedades[self.df_novedades["FICHA"] == self.ficha] 
        if not self.df_activos is None:
            self.df_activos_ficha       = self.df_activos[ self.df_activos["FICHA"] == self.ficha] 
        if not self.df_instructores is None:
            self.df_instructores_ficha  = self.df_instructores[ self.df_instructores["FICHA"] == self.ficha] 

        # 2. creamos el dataframe self.df_datos a partir del dataframe hoja
        columnas_a_copiar = list(COLUMNAS_DATOS.values())[:5]
        nuevas_columnas = {v: None for v in list(COLUMNAS_DATOS.values())[5:25]}
        self.df_datos = self.df_hoja[columnas_a_copiar].copy() 
        self.df_datos = self.df_datos[12:]
        self.df_datos.drop_duplicates(inplace=True)
        self.df_datos = self.df_datos.assign( **nuevas_columnas ) 

        # 3. recorremos el dataframe df_datos y asignamos los valores de las nuevas columnas
        for i in self.df_datos.index: 
            df_aprobado    = self.df_hoja[(self.df_hoja["juicio"] == "APROBADO")    & (self.df_hoja["documento"] == self.df_datos["documento"][i])]
            df_por_evaluar = self.df_hoja[(self.df_hoja["juicio"] == "POR EVALUAR") & (self.df_hoja["documento"] == self.df_datos["documento"][i])]
            df_no_aprobado = self.df_hoja[(self.df_hoja["juicio"] == "NO APROBADO") & (self.df_hoja["documento"] == self.df_datos["documento"][i])]
            self.df_datos.loc[i, "aprobado"]   = len(df_aprobado)
            self.df_datos.loc[i, "porEvaluar"] = len(df_por_evaluar)
            self.df_datos.loc[i, "noAprobado"] = len(df_no_aprobado)
            self.df_datos.loc[i, "orden"]      = numeroDeOrden(self.df_datos["estado"][i], len(df_por_evaluar))
           
            if not self.df_activos_ficha is None and self.df_datos["documento"][i] in self.df_activos_ficha["DOCUMENTO"].values.astype(int):
                self.df_datos.loc[i, "activo"] = "ACTIVO"

            if not self.df_novedades_ficha is None and self.df_datos["estado"][i] == "EN FORMACION":
                for j in self.df_novedades_ficha.index:  
                    if (self.df_datos["documento"][i] == self.df_novedades_ficha["DOCUMENTO"][j]):
                        self.df_datos.loc[i, "enTramite"] = self.df_novedades_ficha["NOVEDAD"][j]
                        break  

            if self.df_datos["estado"][i] == "EN FORMACION":
                contadorTodas = 0
                for competencia in competencias_no_tecnicas:
                    contador = len(df_por_evaluar[df_por_evaluar["competencia"].str.slice(0,5) == competencia[1]])
                    if contador != 0:
                        self.df_datos.loc[i, competencia[0]] = contador
                        contadorTodas += contador 
            
                self.df_datos.loc[i, "TEC"] = len(df_por_evaluar) - contadorTodas

        # Asignar el color a cada fila usando la función color_rows
        for index, row in self.df_datos.iterrows():
            color_value = color_rows(row)[0]
            if isinstance(color_value, str) and color_value.startswith("background-color: "):
                color_value = color_value.replace("background-color: ", "")
            self.df_datos.at[index, "color"] = color_value

        # 4. creamos un dataframe con los nombres de los instructores de cada competencia
        instructores = pd.DataFrame([np.nan] * len(self.df_datos.columns)).transpose()

        instructores['orden'] = 1

        for competencia in [c[0] for c in competencias_no_tecnicas] + ['TEC']:
            try:
                #OJO revisar por que no toma los instructores de datos.xlsx
                if self.df_instructores_ficha[self.df_instructores_ficha['COMPETENCIA'] == competencia]:
                    instructores[competencia] = self.df_instructores_ficha[self.df_instructores_ficha['COMPETENCIA'] == competencia]['INSTRUCTOR'].values
            except:
                instructores[competencia] = getInstructorEnReporte(self.df_hoja, competencia, competencias_no_tecnicas)        

        # 5. insertamos el dataframe instructores en la fila 2
        self.df_datos = pd.concat([self.df_datos.iloc[:0], instructores, self.df_datos.iloc[0:]], ignore_index=True)

    def procesar(self) -> dict:

        columnas = list(COLUMNAS_HOJA.values())
        try:
            self.df_hoja = Entrada().getDataFrame(self.folder, self.archivo, "Hoja", columnas)
            try:
                file_path = os.path.join(self.folder, self.archivo)
                if os.path.isfile(file_path):
                    pass
                    # os.remove(file_path)
            except Exception as e:
                raise Exception(f"Error: no es posible borrar el archivo {e}")
        except Exception as e:
            raise e

        fecha_reporte           = self.df_hoja.iloc[0, 2]
        programa                = self.df_hoja.iloc[4, 2]

        self.ficha              = self.df_hoja.iloc[1, 2] 
        self.codigo_programa    = self.df_hoja.iloc[2, 2] 
        self.version_programa   = self.df_hoja.iloc[3, 2] 
        self.fecha_inicio       = self.df_hoja.iloc[6, 2] 
        self.fecha_fin          = self.df_hoja.iloc[7, 2] 

        print(f"Ficha: {self.ficha} - Programa: {programa}")

        try:
            self.df_hoja.loc[12:, "documento"] = self.df_hoja.loc[12:, "documento"].astype(int)
        except ValueError:
            raise ValueError(f"Error: Hay 'documentos' del reporte de Juicios {ficha} no convertibles a número")

        self.armar_df_datos()
        self.escribirArchivoSalida()

        return {'fecha_reporte'         : fecha_reporte,
                'ficha'                 : self.ficha,
                'programa'              : programa,
                'codigo_programa'       : self.codigo_programa,
                'version_programa'      : self.version_programa,
                'fecha_inicio'          : self.fecha_inicio.strftime("%d-%m-%Y"),
                'fecha_fin'             : self.fecha_fin.strftime("%d-%m-%Y"),
                'df_hoja'               : self.df_hoja,
                'df_datos'              : self.df_datos,
                'df_activos_ficha'      : self.df_activos_ficha,
                'df_novedades_ficha'    : self.df_novedades_ficha,
                'df_instructores_ficha' : self.df_instructores_ficha
                }

if __name__ == "__main__":
    ficha = '3106275'        
    try:
        procesador_jucios = ProcesadorJuicios("upload", f"Reporte de Juicios Evaluativos {ficha}.xls", None, None, None)
        procesador_jucios.procesar()
        print(f"Archivo transformado a XLSX exitosamente para la ficha {ficha}.")
    except Exception as e:
        raise e