
import os
import openpyxl
import pandas as pd

from openpyxl.styles                import Font
from openpyxl.styles                import Alignment
from openpyxl.worksheet.worksheet   import Worksheet

from config                         import FONT_DATOS, SIZE_FONT_DATOS, SIZE_FONT_TITULO, TITULO, TITULO_INSTRUCTORES, TITULO_ULTIMA_FECHA, ALTO_FILA_INSTRUCTORES, ALTO_FILA_FEC_JUICIOS
from config                         import ESTADOS, HOJAS, EXTENSION_EXCEL_365
from procesadorJuiciosHelper        import getLimite_rap_para_normalizar

def color_rows(row, limite_rap_para_normalizar: int):
    color = "White"
    if row['estado'] != "EN FORMACION":
        list_color = [value[1] for value in ESTADOS.values() if value[0] == row['estado']]
        if list_color:
            color = list_color[0]
    else:
        por_evaluar        = row['porEvaluar']
        juicios_productiva = row['PRO']
        se_tramita_novedad = isinstance(row['enTramite'], str)
        if   por_evaluar == 1 and juicios_productiva == 1:
            color = "PaleGreen"
        elif por_evaluar == 1 and juicios_productiva == 0:
            color = "Yellow"
        elif por_evaluar in [n for n in range(2, limite_rap_para_normalizar + 1)]:
            color = "PaleGoldenrod"
        elif se_tramita_novedad:
            color = "DarkRed"
        elif por_evaluar >= limite_rap_para_normalizar and not se_tramita_novedad:
            color = "Red"
        else:
            color = "FireBrick"
    color_final = [f'background-color: {color}']
    return color_final * len(row) 

def ajustarFormatoCeldas(sheet: Worksheet):
    ancho_columnas = HOJAS[sheet.title]['ancho_columnas']
    font = Font(name=FONT_DATOS, size=SIZE_FONT_DATOS)
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = font
        for i in range(len(ancho_columnas)):
            row[i].alignment = Alignment(horizontal='center')      
    for col, width in ancho_columnas:
        sheet.column_dimensions[col].width = width

def write_process_file(folder: str, ficha: str, df_datos: pd.DataFrame, df_novedades_ficha: pd.DataFrame, df_activos_ficha: pd.DataFrame, df_hoja: pd.DataFrame):
    file_name = f"{ficha}.{EXTENSION_EXCEL_365}"
    try:
        with pd.ExcelWriter(os.path.join(folder, file_name), engine="openpyxl") as writer:
            limite_rap_para_normalizar = getLimite_rap_para_normalizar(df_datos)
            df_datos.sort_values(by="orden").style.apply(lambda row: color_rows(row, limite_rap_para_normalizar), axis=1).to_excel(writer, sheet_name='datos', index=False)  
            workbook = writer.book
            worksheet = workbook['datos']
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(25)].hidden = True
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(26)].hidden = True
            ajustarFormatoCeldas(worksheet)
            worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
            worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=10)
            worksheet.cell(2, 1).value = TITULO_INSTRUCTORES
            worksheet.cell(3, 1).value = TITULO_ULTIMA_FECHA            
            worksheet.cell(3, 1).alignment = Alignment(horizontal='right', vertical='center')
            worksheet.cell(2, 1).alignment = Alignment(horizontal='right', vertical='center')
            worksheet.row_dimensions[2].height = ALTO_FILA_INSTRUCTORES
            worksheet.row_dimensions[3].height = ALTO_FILA_FEC_JUICIOS
            for cell in worksheet[2]:
                cell.fill = openpyxl.styles.PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")  
            for cell in worksheet[3]:
                cell.fill = openpyxl.styles.PatternFill(start_color="FFF5E1", end_color="FFF5E1", fill_type="solid")
            for cell in worksheet[2]:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  
            for cell in worksheet[3]:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')  

            if not df_novedades_ficha is None:
                sheet = 'novedades'
                df_novedades_ficha.to_excel(writer, sheet_name=sheet, index=False)
                ajustarFormatoCeldas(workbook[sheet])

            if not df_activos_ficha is None:
                sheet = 'activos'
                df_activos_ficha.to_excel(writer, sheet_name=sheet, index=False)
                ajustarFormatoCeldas(workbook[sheet])

            sheet = 'Hoja'
            df_hoja.to_excel( writer, sheet_name=sheet, index=False)
            worksheet = workbook[sheet]
            ajustarFormatoCeldas(worksheet)
            for fila in range(2, 13):
                worksheet.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
                worksheet.cell(row=fila, column=1).alignment = Alignment(horizontal='left', vertical='center')
                worksheet.cell(row=fila, column=3).alignment = Alignment(horizontal='left', vertical='center')
            for row in worksheet.iter_rows():
                row[5].alignment = Alignment(horizontal='left')  
                row[6].alignment = Alignment(horizontal='left')
            for col in range(1, 12):
                worksheet.cell(1, col).value = None
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
            worksheet.cell(1, 1).value = TITULO
            worksheet.cell(1, 1).font  = Font(name=FONT_DATOS, size=SIZE_FONT_TITULO)
    except Exception as e:
        raise Exception(f"Error: no es posible guardar el archivo: {e}")
